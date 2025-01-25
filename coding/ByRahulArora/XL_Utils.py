import openpyxl


def get_maxRow(filename, SheetName):
    workbook = openpyxl.load_workbook(filename= filename)
    sheet = workbook[SheetName]
    return sheet.max_row

def get_maxColumn(filename, SheetName):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[SheetName]
    return sheet.max_column

def get_cellData(filename, SheetName, row, col):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[SheetName]
    return sheet.cell(row, col).value

def set_cellData(filename, SheetName, row, col, data):
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook[SheetName]
    sheet.cell(row, col).value = data
    workbook.save(filename)


